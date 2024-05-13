import argparse
import csv
import os
from io import StringIO
from itertools import groupby
from operator import itemgetter

import cv2
import numpy as np
import pandas as pd
import pymongo
from PIL import Image
from frameioclient import FrameioClient
from openpyxl.drawing.image import Image as OpenpyxlImage

mongo_client = pymongo.MongoClient("mongodb://localhost:27017/")
database = mongo_client["Crucible"]
baselight_collection = database["Baselight"]
xytech_collection = database["Xytech"]

token = os.environ['FRAMEIOTOKEN']
frame_io_client = FrameioClient(token)


def args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--baselight', type=argparse.FileType('r'), help="import baselight file",
                        required=False)
    parser.add_argument('--xytech', type=argparse.FileType('r'), help="import xytech file",
                        required=False)
    parser.add_argument('--process', type=argparse.FileType('r'), help="process video file",
                        required=False)
    parser.add_argument('--output', action='store_true', help="exports to xlsx",
                        required=False)
    return parser.parse_args()


def handle_args(result):
    if result.baselight is not None:
        import_baselight(result.baselight.name)
    if result.xytech is not None:
        import_xytech(result.xytech.name)
    if result.process is not None:
        process_video(result.process.name, result.output)


def import_baselight(file_path):
    with open(file_path) as f:
        for line in f:
            location_data = line.strip().split()
            location = location_data[0]
            frames = [int(frame) for frame in location_data[1:] if frame.isdigit()]
            data = {'location': location, 'frames': frames}
            baselight_collection.insert_one(data)


def import_xytech(file_path):
    work_order_number = None
    producer = None
    operator = None
    job = None
    notes = None
    locations = []
    with open(file_path) as f:
        for line in f:
            if line.strip().startswith("Xytech Workorder"):
                work_order_number = line.split()[2]
                continue
            if line.strip().startswith("Producer"):
                producer = ' '.join(line.split()[1:])
                continue
            if line.strip().startswith("Operator"):
                operator = ' '.join(line.split()[1:])
                continue
            if line.strip().startswith("Job"):
                job = ' '.join(line.split()[1:])
                continue
            if line.strip().startswith("Notes"):
                notes = next(f).strip()
                continue
            if line.strip().startswith("/"):
                locations.append(line)
        data = [{
            'work_order_number': work_order_number,
            'Producer': producer,
            'Operator': operator,
            'Job': job,
            'Notes': notes,
            'location': locations}]
        xytech_collection.insert_many(data)


def process_video(file_path, output):
    clip = cv2.VideoCapture(file_path)
    fps = clip.get(cv2.CAP_PROP_FPS)
    frame_count = clip.get(cv2.CAP_PROP_FRAME_COUNT)
    if output is not False:
        result = baselight_collection.find({"frames": {"$elemMatch": {"$gte": 0, "$lte": frame_count}}}, {'_id': 0})
    with open('temp_xytech.txt', 'w') as x_f:
        xytech_data = xytech_collection.find()
        for data in xytech_data:
            x_f.write(f"Xytech Workorder {data['work_order_number']}\n\n")
            x_f.write(f"Producer: {data['Producer']}\n")
            x_f.write(f"Operator: {data['Operator']}\n")
            x_f.write(f"Job: {data['Job']}\n\n\n")
            x_f.write("Location:\n")
            for location in data['location']:
                x_f.write(location.strip() + "\n")
            x_f.write("\nNotes:\n")
            x_f.write(data['Notes'] + "\n\n")
    with open('temp_baselight.txt', 'w') as b_f:
        baselight_data = baselight_collection.find()
        for data in baselight_data:
            location = data['location']
            frames = ' '.join(str(frame) for frame in data['frames'])
            b_f.write(location + ' ' + frames + '\n')
    export('temp_xytech.txt', 'temp_baselight.txt')
    os.remove('temp_baselight.txt')
    os.remove('temp_xytech.txt')
    stop_string = 'Location,Frames to fix'
    lines = []
    with open('export.csv', 'r') as file:
        for line in file:
            if line.strip() == stop_string:
                break
            lines.append(line)
    csv_data = ''.join(lines)
    df_existing = pd.read_csv(StringIO(csv_data))
    formatted_frames_data = []
    for item in result:
        frames = sorted(map(int, item['frames']))
        location = item['location']
        for start, end in groupby(enumerate(frames), lambda ix: ix[0] - ix[1]):
            frame_group = list(map(str, (end for _, end in end)))
            if len(frame_group) > 1:
                formatted_frames_data.append(
                    {'location': location, 'Frames to fix': f"{frame_group[0]}-{frame_group[-1]}"})
    df = pd.DataFrame(formatted_frames_data)
    df_combined = pd.concat([df_existing, df], ignore_index=True)
    df_combined['Thumbnail'] = None
    df_combined['Timecode'] = None
    writer = pd.ExcelWriter('excel.xlsx', engine='openpyxl')
    workbook = writer.book
    worksheet = workbook.create_sheet('Sheet1')
    for index, row in df_combined.iterrows():
        frame_range = row['Frames to fix']
        if isinstance(frame_range, float) and np.isnan(frame_range):
            middle_frame = None
        else:
            if isinstance(frame_range, float):
                middle_frame = int(frame_range)
            else:
                start, end = map(int, frame_range.split('-'))
                middle_frame = (start + end) // 2
            thumbnail = generate_thumbnail_for_frame(file_path, middle_frame)
            if thumbnail is not None:
                temp_image_path = f"{middle_frame}.png"
                thumbnail.save(temp_image_path, format="PNG")
                frame_io_client.assets.upload(destination_id="dcf4ff83-ef14-4723-9028-cbef8647506e", filepath=temp_image_path)
                img = OpenpyxlImage(temp_image_path)
                worksheet.add_image(img, f'G{index + 2}')
                df_combined.at[index, 'Timecode'] = get_time_code(fps, middle_frame)
                worksheet.cell(row=index + 2, column=8, value=get_time_code(fps, middle_frame))
    df_combined.to_excel(writer, index=False)
    for idx, col in enumerate(df_combined):
        series = df_combined[col]
        max_len = max((
            series.astype(str).map(len).max(),
            len(str(series.name))
        )) + 1
        worksheet.column_dimensions[worksheet.cell(row=1, column=idx + 1).column_letter].width = max_len
    writer._save()
    os.remove('export.csv')


def generate_thumbnail_for_frame(video_path, frame_number, output_size=(96, 74)):
    clip = cv2.VideoCapture(video_path)
    clip.set(cv2.CAP_PROP_POS_FRAMES, frame_number)
    ret, frame = clip.read()
    clip.release()
    if ret:
        thumbnail = cv2.resize(frame, output_size)
        thumbnail = cv2.cvtColor(thumbnail, cv2.COLOR_BGR2RGB)
        thumbnail = Image.fromarray(thumbnail)
        return thumbnail
    else:
        return None


def convert_to_xls(csv_file, xls_file):
    df = pd.read_csv(csv_file)
    df.to_excel(xls_file, index=False)


def get_time_code(fps, frame):
    total_seconds = frame / fps
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    seconds = int(total_seconds % 60)
    frames = int((total_seconds - int(total_seconds)) * fps)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}:{frames:02d}"


def export(xytech_file_path, baselight_file_path):
    with open("export.csv", 'w', newline='') as csv_file, open(xytech_file_path, 'r') as xytech_file, open(
            baselight_file_path, 'r') as baselight_file:
        writer = csv.writer(csv_file, quoting=csv.QUOTE_NONE)
        writer.writerow(['Producer', 'Operator', 'Job', 'Notes'])
        locations = {}
        job_details = {h: '' for h in ['Producer', 'Operator', 'Job', 'Notes']}
        for line in xytech_file:
            if line.startswith('/'):
                parts = line.strip().split("/", 3)[1:]
                locations[parts[2]] = parts[:2]
            elif any(line.startswith(h) for h in job_details):
                key, value = line.strip().split(':', 1)
                job_details[key] = value.strip() if key != 'Notes' else next(xytech_file, '').strip()
        writer.writerows(
            [[job_details[h] for h in ['Producer', 'Operator', 'Job', 'Notes']], [], ["Location", "Frames to fix"]])
        for line in baselight_file:
            location = line.strip().split()
            location_key = location[0].split("/", 2)[2:][0]
            if location_key in locations:
                formatted_frames = []
                for k, g in groupby(
                        enumerate([item for item in location[1:] if '<err>' not in item and '<null>' not in item]),
                        lambda ix: int(ix[0]) - int(ix[1])):
                    frames = list(map(itemgetter(1), g))
                    formatted_frames.append(frames[0] + "-" + frames[len(frames) - 1] if len(frames) > 1 else frames[0])
                base_location = "/" + "/".join(locations[location_key][:2]) + "/" + location_key
                writer.writerows([[f"{base_location} {frame}"] for frame in formatted_frames])


if __name__ == '__main__':
    handle_args(args())
